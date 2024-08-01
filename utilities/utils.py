import softest
import logging
import inspect
from openpyxl import workbook, load_workbook


class Utils(softest.TestCase):
    def custlogger(loglevel=logging.DEBUG):
        # set class/method name from where its called
        logger_name = inspect.stack()[1][3]
        # create logger
        logger = logging.getLogger(logger_name)
        logger.setLevel(loglevel)
        # create console handler or file handler and set the log level
        fh = logging.FileHandler("automation.log", mode='a')
        # create formatter - how you wnat your logs to be formatted
        formatter = logging.Formatter('%(created)f - %(asctime)s - %(levelname)s - %(levelno)s : %(message)s',
                    datefmt='%m/%d/%Y %I:%M:%S %p')
        # add formatter to console or file handler
        fh.setFormatter(formatter)
        # add console handler to logger
        logger.addHandler(fh)
        return logger


    # This code will read data from excel file
    def readDataFromExcel(file_name, sheet_name):
        # Create the empty list to store the value
        datalist = []
        # Create object for loading the workbook
        wb = load_workbook(filename=file_name)
        # Create object for sheet
        sh = wb[sheet_name]
        # To get the max row value
        row_ct = sh.max_row
        # To get the max col value
        col_ct = sh.max_column
        # loop to iterate the rows and column data
        for i in range(2, row_ct + 1):
            # This will be storing data for particular row
            row = []
            for j in range(1, col_ct + 1):
                row.append(sh.cell(row=i, column=j).value)
            datalist.append(row)
        return datalist


    # This code will read data from CSV
    # def readdatafromcsv(filename):
    #     # Create an empty list
    #     datalist = []
    #     # Open CSV file
    #     csvdata = open(filename, "r")
    #     # Create CSV reader
    #     reader = csv.reader(csvdata)
    #     # skip header
    #     next(reader)
    #     # Add CSV rows to list
    #     for rows in reader:
    #         datalist.append(rows)
    #     return datalist






